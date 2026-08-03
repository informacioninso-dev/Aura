"""
Respaldo completo de la cuenta en un solo XLSX multi-hoja (export + import).

A diferencia de `importar.py` (pensado para estados de cuenta bancarios en una
tabla plana), aqui el objetivo es **mover la cuenta entera**: todas las entidades
con datos financieros, en un archivo que se puede volver a importar en otra
cuenta. Una hoja por tipo:

  Movimientos          -> Ingreso, IngresoPuntual, GastoCorriente, GastoNoCorriente
  Consumos_variables   -> GastoCorrienteEjecucion (gasto real por mes de un rubro)
  Diferidos            -> Diferido (creditos a cuotas)
  Cuentas              -> CuentaPorCobrar (te deben / debes)
  Categorias           -> Categoria (las propias del usuario)

El export escribe fechas como texto ISO y montos como numero; el import las
vuelve a parsear con los mismos helpers que `importar.py`.
"""

import io

from django.db import transaction

from .importar import _parse_fecha, _parse_monto, _normalizar, FRECUENCIAS_VALIDAS

# Tope defensivo por hoja (el modulo esta gated a usuarios 'asesor', pero igual
# no queremos que un archivo enorme tumbe el proceso).
MAX_FILAS_HOJA = 10000

MOVIMIENTOS_HEADERS = [
    'tipo', 'descripcion', 'monto', 'categoria', 'frecuencia', 'tipo_monto',
    'fecha_inicio', 'fecha_fin', 'fecha', 'activo', 'notas',
]
CONSUMOS_HEADERS = ['rubro_descripcion', 'rubro_categoria', 'anio', 'mes', 'fecha', 'descripcion', 'monto_real']
DIFERIDOS_HEADERS = ['descripcion', 'categoria', 'monto_total', 'num_cuotas', 'cuota_mensual', 'fecha_inicio', 'fecha_fin', 'activo']
CUENTAS_HEADERS = ['direccion', 'persona', 'concepto', 'monto_total', 'monto_cobrado', 'fecha_prestamo', 'fecha_recordatorio', 'notas']
CATEGORIAS_HEADERS = ['nombre', 'icono', 'limite_mensual']

TIPOS_MOVIMIENTO = {'ingreso_fijo', 'ingreso_puntual', 'gasto_fijo', 'gasto_variable', 'gasto_puntual'}


# --------------------------------------------------------------------------- #
# Helpers de serializacion / parseo
# --------------------------------------------------------------------------- #
def _s(value):
    """Fecha/valor -> texto plano para la celda."""
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value)


def _num(value):
    """Decimal -> float para que openpyxl lo escriba como numero."""
    return float(value) if value is not None else 0.0


def _bool_cell(value):
    return 'si' if value else 'no'


def _parse_bool(raw, default=True):
    v = _normalizar(raw)
    if v in ('si', 'strue', '1', 'yes', 'x', 'verdadero'):
        return True
    if v in ('no', 'false', '0', 'inactivo', 'falso'):
        return False
    return default


def _parse_int(raw):
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- #
# EXPORT
# --------------------------------------------------------------------------- #
def exportar_cuenta_xlsx(usuario) -> bytes:
    import openpyxl

    from .models import (
        Categoria, Ingreso, IngresoPuntual, GastoCorriente,
        GastoCorrienteEjecucion, GastoNoCorriente, Diferido, CuentaPorCobrar,
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Movimientos')
    ws.append(MOVIMIENTOS_HEADERS)
    for i in Ingreso.objects.filter(usuario=usuario):
        ws.append(['ingreso_fijo', i.descripcion, _num(i.monto), '', i.frecuencia, '', _s(i.fecha_inicio), _s(i.fecha_fin), '', _bool_cell(i.activo), ''])
    for i in IngresoPuntual.objects.filter(usuario=usuario):
        ws.append(['ingreso_puntual', i.descripcion, _num(i.monto), '', '', '', '', '', _s(i.fecha), '', i.notas or ''])
    for g in GastoCorriente.objects.filter(usuario=usuario):
        tipo = 'gasto_variable' if g.tipo_monto == 'variable' else 'gasto_fijo'
        ws.append([tipo, g.descripcion, _num(g.monto), g.categoria, g.frecuencia, g.tipo_monto, _s(g.fecha_inicio), _s(g.fecha_fin), '', _bool_cell(g.activo), ''])
    for g in GastoNoCorriente.objects.filter(usuario=usuario):
        ws.append(['gasto_puntual', g.descripcion, _num(g.monto), g.categoria, '', '', '', '', _s(g.fecha), '', g.notas or ''])

    ws = wb.create_sheet('Consumos_variables')
    ws.append(CONSUMOS_HEADERS)
    for e in GastoCorrienteEjecucion.objects.filter(gasto__usuario=usuario).select_related('gasto'):
        ws.append([e.gasto.descripcion, e.gasto.categoria, e.anio, e.mes, _s(e.fecha), e.descripcion or '', _num(e.monto_real)])

    ws = wb.create_sheet('Diferidos')
    ws.append(DIFERIDOS_HEADERS)
    for d in Diferido.objects.filter(usuario=usuario):
        ws.append([d.descripcion, d.categoria, _num(d.monto_total), d.num_cuotas, _num(d.cuota_mensual), _s(d.fecha_inicio), _s(d.fecha_fin), _bool_cell(d.activo)])

    ws = wb.create_sheet('Cuentas')
    ws.append(CUENTAS_HEADERS)
    for c in CuentaPorCobrar.objects.filter(usuario=usuario):
        ws.append([c.direccion, c.persona, c.concepto, _num(c.monto_total), _num(c.monto_cobrado), _s(c.fecha_prestamo), _s(c.fecha_recordatorio), c.notas or ''])

    ws = wb.create_sheet('Categorias')
    ws.append(CATEGORIAS_HEADERS)
    for cat in Categoria.objects.filter(usuario=usuario):
        ws.append([cat.nombre, cat.icono, _num(cat.limite_mensual) if cat.limite_mensual is not None else ''])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# IMPORT
# --------------------------------------------------------------------------- #
def _leer_hoja(wb, titulo, headers):
    """Devuelve lista de dicts {header: valor} de la hoja, o [] si no existe."""
    if titulo not in wb.sheetnames:
        return []
    ws = wb[titulo]
    filas = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            cabecera = [_normalizar(c) for c in row]
            continue
        if idx > MAX_FILAS_HOJA:
            raise ValueError(f'La hoja "{titulo}" supera el maximo de {MAX_FILAS_HOJA} filas.')
        if not any(str(c).strip() for c in row if c is not None):
            continue
        d = {}
        for h in headers:
            try:
                d[h] = row[cabecera.index(h)]
            except (ValueError, IndexError):
                d[h] = None
        filas.append(d)
    return filas


def importar_cuenta_xlsx(usuario, file_bytes: bytes) -> dict:
    """
    Crea (de forma aditiva) todas las entidades del respaldo. Devuelve el conteo
    por tipo y una lista de avisos por filas omitidas. Todo en una transaccion.
    """
    import openpyxl

    from .models import (
        Categoria, Ingreso, IngresoPuntual, GastoCorriente,
        GastoCorrienteEjecucion, GastoNoCorriente, Diferido, CuentaPorCobrar,
        TIPO_MONTO_VARIABLE, TIPO_MONTO_FIJO,
    )
    from .importar import _validar_contenedor_xlsx
    from .utils import invalidate_finanzas_cache, sincronizar_inicio_rubro

    _validar_contenedor_xlsx(file_bytes)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('El archivo XLSX esta danado o no es valido.') from exc

    try:
        movimientos = _leer_hoja(wb, 'Movimientos', MOVIMIENTOS_HEADERS)
        consumos = _leer_hoja(wb, 'Consumos_variables', CONSUMOS_HEADERS)
        diferidos = _leer_hoja(wb, 'Diferidos', DIFERIDOS_HEADERS)
        cuentas = _leer_hoja(wb, 'Cuentas', CUENTAS_HEADERS)
        categorias = _leer_hoja(wb, 'Categorias', CATEGORIAS_HEADERS)
    finally:
        wb.close()

    if not any([movimientos, consumos, diferidos, cuentas, categorias]):
        raise ValueError('El archivo no tiene ninguna hoja reconocible (Movimientos, Diferidos, Cuentas, Categorias, Consumos_variables).')

    avisos = []
    conteo = {'categorias': 0, 'ingresos': 0, 'gastos': 0, 'consumos': 0, 'diferidos': 0, 'cuentas': 0}
    fechas_afectadas = []

    with transaction.atomic():
        # 1) Categorias primero (los movimientos las referencian por nombre).
        for row in categorias:
            nombre = str(row.get('nombre') or '').strip()[:50]
            if not nombre:
                continue
            limite = _parse_monto(row.get('limite_mensual')) if str(row.get('limite_mensual') or '').strip() else None
            _, creada = Categoria.objects.get_or_create(
                usuario=usuario, nombre=nombre,
                defaults={'icono': (str(row.get('icono') or '📦').strip()[:10] or '📦'), 'limite_mensual': limite},
            )
            if creada:
                conteo['categorias'] += 1

        # 2) Movimientos.
        ingresos_fijos, ingresos_puntuales, gastos_fijos, gastos_puntuales = [], [], [], []
        rubros_variables = []  # (descripcion_norm, categoria_norm, instancia)

        for row in movimientos:
            tipo = _normalizar(row.get('tipo'))
            if tipo not in TIPOS_MOVIMIENTO:
                avisos.append(f'Movimiento con tipo desconocido: "{row.get("tipo")}" (omitido)')
                continue
            desc = (str(row.get('descripcion') or '').strip() or '(sin descripcion)')[:200]
            monto = _parse_monto(row.get('monto'))
            # Los rubros variables se crean con monto 0 (es solo un estimado); el
            # resto exige monto > 0.
            if monto is None or monto < 0 or (monto == 0 and tipo != 'gasto_variable'):
                avisos.append(f'Movimiento "{desc}" con monto invalido (omitido)')
                continue
            cat = (str(row.get('categoria') or 'otro').strip().lower() or 'otro')[:50]
            notas = str(row.get('notas') or '')
            activo = _parse_bool(row.get('activo'), default=True)

            if tipo in ('ingreso_fijo', 'gasto_fijo', 'gasto_variable'):
                fi = _parse_fecha(str(row.get('fecha_inicio') or '').strip())
                if not fi:
                    avisos.append(f'Recurrente "{desc}" sin fecha_inicio valida (omitido)')
                    continue
                ff = _parse_fecha(str(row.get('fecha_fin') or '').strip()) or None
                frecuencia = _normalizar(row.get('frecuencia'))
                if frecuencia not in FRECUENCIAS_VALIDAS:
                    frecuencia = 'mensual'
                fechas_afectadas.append(fi)
                if tipo == 'ingreso_fijo':
                    ingresos_fijos.append(Ingreso(usuario=usuario, descripcion=desc, monto=monto, frecuencia=frecuencia, fecha_inicio=fi, fecha_fin=ff, activo=activo))
                else:
                    tm = TIPO_MONTO_VARIABLE if tipo == 'gasto_variable' else TIPO_MONTO_FIJO
                    inst = GastoCorriente(usuario=usuario, descripcion=desc, categoria=cat, monto=monto, tipo_monto=tm, frecuencia=frecuencia, fecha_inicio=fi, fecha_fin=ff, activo=activo)
                    if tipo == 'gasto_variable':
                        rubros_variables.append((_normalizar(desc), _normalizar(cat), inst))
                    gastos_fijos.append(inst)
            else:  # puntuales
                f = _parse_fecha(str(row.get('fecha') or '').strip())
                if not f:
                    avisos.append(f'Puntual "{desc}" sin fecha valida (omitido)')
                    continue
                fechas_afectadas.append(f)
                if tipo == 'ingreso_puntual':
                    ingresos_puntuales.append(IngresoPuntual(usuario=usuario, descripcion=desc, monto=monto, fecha=f, notas=notas))
                else:
                    gastos_puntuales.append(GastoNoCorriente(usuario=usuario, descripcion=desc, categoria=cat, monto=monto, fecha=f, notas=notas))

        if ingresos_fijos:
            Ingreso.objects.bulk_create(ingresos_fijos, batch_size=500)
        if ingresos_puntuales:
            IngresoPuntual.objects.bulk_create(ingresos_puntuales, batch_size=500)
        if gastos_fijos:
            GastoCorriente.objects.bulk_create(gastos_fijos, batch_size=500)
        if gastos_puntuales:
            GastoNoCorriente.objects.bulk_create(gastos_puntuales, batch_size=500)
        conteo['ingresos'] = len(ingresos_fijos) + len(ingresos_puntuales)
        conteo['gastos'] = len(gastos_fijos) + len(gastos_puntuales)

        # 3) Consumos de variables -> referencian un rubro por descripcion+categoria.
        #    Se reconsulta para tener los pk de los rubros recien creados.
        mapa_rubros = {}
        for gc in GastoCorriente.objects.filter(usuario=usuario, tipo_monto=TIPO_MONTO_VARIABLE):
            mapa_rubros[(_normalizar(gc.descripcion), _normalizar(gc.categoria))] = gc.id
        ejecuciones = []
        gastos_con_consumo = set()
        for row in consumos:
            clave = (_normalizar(row.get('rubro_descripcion')), _normalizar(row.get('rubro_categoria')))
            gasto_id = mapa_rubros.get(clave)
            if not gasto_id:
                avisos.append(f'Consumo sin rubro variable coincidente: "{row.get("rubro_descripcion")}" (omitido)')
                continue
            monto = _parse_monto(row.get('monto_real'))
            anio = _parse_int(row.get('anio'))
            mes = _parse_int(row.get('mes'))
            if monto is None or monto <= 0 or not anio or not mes or not (1 <= mes <= 12):
                avisos.append(f'Consumo de "{row.get("rubro_descripcion")}" con datos invalidos (omitido)')
                continue
            fecha = _parse_fecha(str(row.get('fecha') or '').strip()) or None
            ejecuciones.append(GastoCorrienteEjecucion(
                gasto_id=gasto_id, anio=anio, mes=mes, fecha=fecha,
                descripcion=(str(row.get('descripcion') or '').strip())[:200], monto_real=monto,
            ))
            gastos_con_consumo.add(gasto_id)
            if fecha:
                fechas_afectadas.append(fecha)
        if ejecuciones:
            GastoCorrienteEjecucion.objects.bulk_create(ejecuciones, batch_size=500)
            # bulk_create no dispara save(): sincronizamos el inicio de cada rubro.
            for gid in gastos_con_consumo:
                sincronizar_inicio_rubro(gid)
            conteo['consumos'] = len(ejecuciones)

        # 4) Diferidos.
        difs = []
        for row in diferidos:
            desc = (str(row.get('descripcion') or '').strip())[:200]
            total = _parse_monto(row.get('monto_total'))
            cuota = _parse_monto(row.get('cuota_mensual'))
            ncuotas = _parse_int(row.get('num_cuotas'))
            fi = _parse_fecha(str(row.get('fecha_inicio') or '').strip())
            ff = _parse_fecha(str(row.get('fecha_fin') or '').strip())
            if not desc or total is None or cuota is None or not ncuotas or not fi or not ff:
                avisos.append(f'Diferido "{desc or "?"}" con datos incompletos (omitido)')
                continue
            difs.append(Diferido(
                usuario=usuario, descripcion=desc,
                categoria=(str(row.get('categoria') or 'otro').strip().lower() or 'otro')[:50],
                monto_total=total, num_cuotas=ncuotas, cuota_mensual=cuota,
                fecha_inicio=fi, fecha_fin=ff, activo=_parse_bool(row.get('activo'), default=True),
            ))
            fechas_afectadas.append(fi)
        if difs:
            Diferido.objects.bulk_create(difs, batch_size=500)
            conteo['diferidos'] = len(difs)

        # 5) Cuentas por cobrar / pagar.
        cxc = []
        for row in cuentas:
            persona = (str(row.get('persona') or '').strip())[:120]
            total = _parse_monto(row.get('monto_total'))
            fp = _parse_fecha(str(row.get('fecha_prestamo') or '').strip())
            if not persona or total is None or not fp:
                avisos.append(f'Cuenta de "{persona or "?"}" con datos incompletos (omitida)')
                continue
            direccion = _normalizar(row.get('direccion'))
            if direccion not in ('me_deben', 'debo'):
                direccion = 'me_deben'
            cobrado = _parse_monto(row.get('monto_cobrado')) or 0
            cxc.append(CuentaPorCobrar(
                usuario=usuario, direccion=direccion, persona=persona,
                concepto=(str(row.get('concepto') or '').strip())[:200],
                monto_total=total, monto_cobrado=cobrado, fecha_prestamo=fp,
                fecha_recordatorio=_parse_fecha(str(row.get('fecha_recordatorio') or '').strip()) or None,
                notas=str(row.get('notas') or ''),
            ))
        if cxc:
            CuentaPorCobrar.objects.bulk_create(cxc, batch_size=500)
            conteo['cuentas'] = len(cxc)

        if fechas_afectadas:
            invalidate_finanzas_cache(usuario.pk, min(fechas_afectadas))

    return {'conteo': conteo, 'avisos': avisos, 'total': sum(conteo.values())}
